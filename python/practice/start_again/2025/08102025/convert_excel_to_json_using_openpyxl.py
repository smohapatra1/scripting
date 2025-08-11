#   https://www.geeksforgeeks.org/python/convert-excel-to-json-with-python/

from openpyxl import load_workbook
from json import dumps

wb = load_workbook("test.xls")
sheet = wb["test"]
rows = sheet.max_row
columns = sheet.max_column
lst = []
for i in range(1, rows):
    row = {}
    for j in range(1, columns):
        column_name = sheet.cell(row=1, column=j)
        row_data = sheet.cell(row=i+1, column=j)
        row.update(
            {
                column_name.value: row_data.value
            }
        )
    list.append(row)
json_data = dumps(lst)
print (json_data)
